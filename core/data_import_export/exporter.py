"""
🌸 若曦V2 - 健康数据导出器
支持多种格式的数据导出
"""
from typing import Dict, List, Optional, BinaryIO, Union
from dataclasses import dataclass
from datetime import datetime, timedelta
from enum import Enum
from io import BytesIO
import csv
import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExportFormat(Enum):
    """导出格式"""
    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"
    PDF = "pdf"


class ExportTemplate(Enum):
    """导出模板"""
    RAW_DATA = "raw_data"           # 原始数据
    DAILY_SUMMARY = "daily_summary"  # 每日汇总
    WEEKLY_REPORT = "weekly_report" # 周报
    MONTHLY_REPORT = "monthly_report"# 月报
    FOR_DOCTOR = "for_doctor"       # 医生格式


@dataclass
class ExportResult:
    """导出结果"""
    success: bool
    file_content: Optional[bytes]
    file_name: str
    mime_type: str
    record_count: int
    date_range: Dict[str, str]
    warnings: List[str]
    
    def to_dict(self) -> Dict:
        return {
            "success": self.success,
            "file_name": self.file_name,
            "mime_type": self.mime_type,
            "record_count": self.record_count,
            "date_range": self.date_range,
            "warnings": self.warnings
        }


class DataExporter:
    """
    健康数据导出器
    
    支持:
    - CSV/Excel/JSON/PDF 格式
    - 多种导出模板 (原始/日报/周报/月报)
    - 医生专用格式
    - 数据可视化图表
    - 自动邮件发送
    """
    
    def __init__(self):
        self._templates = {}
    
    async def export(
        self,
        data: List[Dict],
        data_type: str,
        user_id: str,
        export_format: ExportFormat = ExportFormat.EXCEL,
        template: ExportTemplate = ExportTemplate.RAW_DATA,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        include_charts: bool = False
    ) -> ExportResult:
        """
        导出数据
        
        Args:
            data: 原始数据列表
            data_type: 数据类型
            user_id: 用户ID
            export_format: 导出格式
            template: 导出模板
            date_from: 开始日期
            date_to: 结束日期
            include_charts: 是否包含图表
        """
        if not data:
            return ExportResult(
                success=False,
                file_content=None,
                file_name="",
                mime_type="",
                record_count=0,
                date_range={},
                warnings=["没有数据可导出"]
            )
        
        # 准备数据
        df = pd.DataFrame(data)
        
        # 应用模板
        df, metadata = self._apply_template(df, template, data_type)
        
        # 生成文件
        file_content, mime_type, extension = self._generate_file(
            df, export_format, include_charts
        )
        
        # 生成文件名
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        file_name = f"{data_type}_{template.value}_{timestamp}.{extension}"
        
        # 计算日期范围
        date_range = self._calculate_date_range(data)
        
        return ExportResult(
            success=True,
            file_content=file_content,
            file_name=file_name,
            mime_type=mime_type,
            record_count=len(df),
            date_range=date_range,
            warnings=metadata.get('warnings', [])
        )
    
    async def export_health_report(
        self,
        user_id: str,
        report_type: str = "weekly",
        export_format: ExportFormat = ExportFormat.PDF
    ) -> ExportResult:
        """
        导出健康报告
        
        生成综合健康报告，包含多种数据类型
        """
        warnings = []
        
        # 收集用户健康数据
        # TODO: 从数据库获取真实数据
        
        # 模拟数据
        mock_data = {
            "blood_pressure": [],
            "blood_glucose": [],
            "weight": [],
            "sleep": [],
            "heart_rate": [],
            "steps": []
        }
        
        # 生成报告
        if export_format == ExportFormat.PDF:
            # TODO: 实现PDF报告生成
            return ExportResult(
                success=False,
                file_content=None,
                file_name="",
                mime_type="",
                record_count=0,
                date_range={},
                warnings=["PDF生成功能开发中"]
            )
        
        else:
            # 使用Excel格式作为报告
            return await self.export(
                data=[{"section": "健康报告", "type": report_type}],
                data_type="health_report",
                user_id=user_id,
                export_format=export_format,
                template=ExportTemplate.WEEKLY_REPORT if report_type == "weekly" else ExportTemplate.MONTHLY_REPORT
            )
    
    async def export_batch(
        self,
        data_types: List[str],
        user_id: str,
        export_format: ExportFormat = ExportFormat.EXCEL,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None
    ) -> Dict[str, ExportResult]:
        """
        批量导出多种数据类型
        
        返回多个ExportResult的字典
        """
        results = {}
        
        for data_type in data_types:
            # TODO: 获取真实数据
            mock_data = []
            
            result = await self.export(
                data=mock_data,
                data_type=data_type,
                user_id=user_id,
                export_format=export_format,
                template=ExportTemplate.RAW_DATA,
                date_from=date_from,
                date_to=date_to
            )
            
            results[data_type] = result
        
        return results
    
    def _apply_template(
        self,
        df: pd.DataFrame,
        template: ExportTemplate,
        data_type: str
    ) -> tuple[pd.DataFrame, Dict]:
        """应用导出模板"""
        metadata = {'warnings': []}
        
        if template == ExportTemplate.RAW_DATA:
            # 原始数据，不做处理
            return df, metadata
        
        elif template == ExportTemplate.DAILY_SUMMARY:
            # 每日汇总
            if 'timestamp' in df.columns:
                df['date'] = pd.to_datetime(df['timestamp']).dt.date
                
                if data_type == 'blood_pressure':
                    df = df.groupby('date').agg({
                        'systolic': ['mean', 'min', 'max', 'count'],
                        'diastolic': ['mean', 'min', 'max']
                    }).reset_index()
                    df.columns = ['date', 'systolic_avg', 'systolic_min', 'systolic_max', 'readings', 
                                 'diastolic_avg', 'diastolic_min', 'diastolic_max']
                
                elif data_type == 'blood_glucose':
                    df = df.groupby('date').agg({
                        'value': ['mean', 'min', 'max', 'count']
                    }).reset_index()
                    df.columns = ['date', 'glucose_avg', 'glucose_min', 'glucose_max', 'readings']
        
        elif template == ExportTemplate.WEEKLY_REPORT:
            # 周报格式
            metadata['warnings'].append("周报模板包含周平均值和趋势分析")
            # 添加周统计信息
            if 'date' in df.columns or 'timestamp' in df.columns:
                date_col = 'date' if 'date' in df.columns else 'timestamp'
                df[date_col] = pd.to_datetime(df[date_col])
                df['week'] = df[date_col].dt.isocalendar().week
                df['year'] = df[date_col].dt.isocalendar().year
        
        elif template == ExportTemplate.FOR_DOCTOR:
            # 医生格式：标准、专业
            # 添加标准化单位
            metadata['warnings'].append("医生格式使用国际标准单位")
            
            # 添加统计信息
            stats_row = pd.DataFrame([{
                '统计项': f'{data_type}统计',
                '记录数': len(df),
                '首次记录': df['timestamp'].min() if 'timestamp' in df.columns else 'N/A',
                '最新记录': df['timestamp'].max() if 'timestamp' in df.columns else 'N/A'
            }])
            
            # 将统计数据添加到开头
            df = pd.concat([stats_row, df], ignore_index=True)
        
        return df, metadata
    
    def _generate_file(
        self,
        df: pd.DataFrame,
        export_format: ExportFormat,
        include_charts: bool
    ) -> tuple[bytes, str, str]:
        """生成文件内容"""
        
        if export_format == ExportFormat.CSV:
            output = BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            return output.getvalue(), 'text/csv', 'csv'
        
        elif export_format == ExportFormat.EXCEL:
            return self._generate_excel(df, include_charts)
        
        elif export_format == ExportFormat.JSON:
            output = BytesIO()
            json.dump(df.to_dict('records'), output, ensure_ascii=False, indent=2)
            return output.getvalue(), 'application/json', 'json'
        
        elif export_format == ExportFormat.PDF:
            # TODO: 实现PDF生成
            raise NotImplementedError("PDF导出功能开发中")
        
        else:
            raise ValueError(f"不支持的导出格式: {export_format}")
    
    def _generate_excel(
        self,
        df: pd.DataFrame,
        include_charts: bool
    ) -> tuple[bytes, str, str]:
        """生成Excel文件"""
        output = BytesIO()
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "数据"
        
        # 样式
        header_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        header_font = Font(bold=True, color="2C5F7C")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                
                # 表头样式
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(vertical='center')
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结表头
        ws.freeze_panes = 'A2'
        
        # TODO: 如果include_charts为True，添加图表sheet
        
        # 保存
        wb.save(output)
        return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx'
    
    def _calculate_date_range(self, data: List[Dict]) -> Dict[str, str]:
        """计算数据日期范围"""
        if not data:
            return {}
        
        timestamps = []
        for record in data:
            ts = record.get('timestamp') or record.get('date')
            if ts:
                timestamps.append(ts)
        
        if not timestamps:
            return {}
        
        # 解析日期
        dates = []
        for ts in timestamps:
            try:
                if isinstance(ts, str):
                    dates.append(pd.to_datetime(ts))
                elif isinstance(ts, datetime):
                    dates.append(ts)
            except:
                pass
        
        if dates:
            return {
                'from': min(dates).strftime('%Y-%m-%d'),
                'to': max(dates).strftime('%Y-%m-%d')
            }
        
        return {}
    
    async def send_by_email(
        self,
        export_result: ExportResult,
        email: str,
        subject: Optional[str] = None
    ) -> bool:
        """
        通过邮件发送导出文件
        
        TODO: 集成邮件服务
        """
        if not export_result.success:
            return False
        
        print(f"📧 发送导出文件到 {email}")
        print(f"📎 附件: {export_result.file_name}")
        
        return True


# 导出器实例
data_exporter = DataExporter()
